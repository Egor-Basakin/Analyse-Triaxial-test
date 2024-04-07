import tkinter as tk
from tkinter import filedialog
import pandas as pd
import numpy as np
from scipy.signal import find_peaks
from scipy.optimize import curve_fit
from tkinter import ttk
from tkinter import simpledialog
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import csv
import os
import shutil
import threading
import time
from openpyxl import load_workbook
from openpyxl.chart import ScatterChart, Reference, Series
import math
from scipy import interpolate
from scipy.interpolate import interp1d
import asyncio

df = None
df_filtr = None
csv_files = []
checkboxes = []
checkbox_frame = None  # Определяем переменную checkbox_frame глобально
height = 100
# Функции для обработки событий


# Функция для получения списка файлов CSV в директории
def list_csv_files(directory):
    return [file for file in os.listdir(directory) if file.endswith('.csv')]

# Функция для обновления списка файлов
def update_file_list():
    global csv_files
    csv_files = list_csv_files(directory_path)  # Обновляем список файлов
    file_listbox.delete(0, tk.END)  # Очищаем список в GUI
    for file in csv_files:
        file_listbox.insert(tk.END, file)  # Добавляем файлы в список в GUI

def on_file_select(event):
    global df
    selected_file = file_listbox.get(file_listbox.curselection())
    full_path = os.path.join(directory_path, selected_file)  # Формирование полного пути к файлу
    print(f"Выбран файл: {full_path}")
    Name_file.config(text=f"File: {full_path}")
     # Загрузка данных из файла
    try:
        df = pd.read_csv(full_path)  # или измените на соответствующий формат файла, если не CSV
        create_checkboxes()

        plot_base() 


    except Exception as e:
        print("Ошибка при загрузке файла:", e)






def plot_base():
    global df
    x = df['Axial Displacement (mm)'] #Присвоение x 
    y = df['Deviator Stress (kPa)'] # Присвоение у

    # Построение графика
   # Постройте график
    plt.figure()
    plt.plot(x, y, label='Данные')

    plt.legend()

    # Добавьте заголовок и подписи к осям
    plt.title('Базовый график')
    plt.xlabel('Х')
    plt.ylabel('Y')
    # Добавление сетки
    plt.grid(True)
    # Отобразите график
    plt.show()

def create_checkboxes():
    global df, checkboxes, checkbox_frame, canvas
    
    # Проверяем, загружен ли датафрейм и есть ли столбец 'Stage Number'
    if df is not None and 'Stage Number' in df.columns:
        if checkbox_frame:
            checkbox_frame.destroy()  # Уничтожаем предыдущий фрейм с чекбоксами, если он существует
        
        checkbox_frame = tk.Frame(canvas)  # Создаем новый фрейм для чекбоксов внутри холста
        canvas.create_window((0, 0), window=checkbox_frame, anchor='nw')  # Привязываем фрейм к холсту
        
        checkboxes = []  # Обнуляем список чекбоксов
        
        # Создаем чекбоксы для каждой уникальной стадии
        for stage_number in df['Stage Number'].unique():
            var = tk.IntVar()
            checkbox = tk.Checkbutton(checkbox_frame, text=f"Стадия {stage_number}", variable=var)
            checkbox.pack(anchor='w')
            checkboxes.append(var)  # Сохраняем переменные в список
    else:
        print("DataFrame не содержит столбец 'Stage Number' или не загружен.")




def save_filtered_data():
    global df, df_filtr
    if df is not None:
        selected_stages = []
        for i, var in enumerate(checkboxes):
            if var.get() == 1:
                selected_stages.append(i + 1)  # Индексация начинается с 0, поэтому добавляем 1 для получения номера стадии
                
        df_filtr = df[df['Stage Number'].isin(selected_stages)]
        plot_filter()
    else:
        print("Нет данных для фильтрации.")


def plot_filter():
    global df_filtr
    x = df_filtr['Axial Displacement (mm)'] #Присвоение x 
    y = df_filtr['Deviator Stress (kPa)'] # Присвоение у

    # Построение графика
   # Постройте график
    plt.figure()
    plt.plot(x, y, label='Данные')

    plt.legend()

    # Добавьте заголовок и подписи к осям
    plt.title('отфильтрованные график по стадиям')
    plt.xlabel('Х')
    plt.ylabel('Y')
    # Добавление сетки
    plt.grid(True)
    # Отобразите график
    plt.show()

def reset_displacement():
    global df_filtr, height
    
    if df_filtr is not None:
          
        min_displacement = df_filtr['Axial Displacement (mm)'].iloc[0]  # Находим минимальное значение перемещения
        if min_displacement == 0:
                print("Высота образца не требует вычислений, так как минимальное перемещение равно нулю.")
        else:
            # Обнуляем значения перемещения, вычитая минимальное значение
            df_filtr['Axial Displacement (mm)'] -= min_displacement
            # Вычисление высоты образца (height)
            height = 100 - min_displacement
            print("Высота образца зафиксирована", height)
        plot_filter()
    else:
        print("Нет данных для обнуления перемещения.")



def reset_deviator():
    global df_filtr
    
    if df_filtr is not None:
        min_deviator = df_filtr['Deviator Stress (kPa)'].iloc[0]  # Находим минимальное значение девиатора
        df_filtr['Deviator Stress (kPa)'].replace(min_deviator, 0, inplace=True)  # Заменяем минимальное значение на 0
        #проделываем тоже самое для столбца load cell чтобы при учете поправки выводились норм данные 
        min_load = df_filtr['Load Cell (kN)'].iloc[0]  
        df_filtr['Load Cell (kN)'].replace(min_load, 0, inplace=True)  
        print("Минимальное значение девиатора заменено на 0.")
        plot_filter()
    else:
        print("Нет данных для обнуления значения девиатора.")


def select_from_list():
    global checkboxes
    for var in checkboxes:
        var.set(1)  # Устанавливаем состояние чекбоксов из списка checkboxes в "выбрано"

def deselect_from_list():
    global checkboxes
    for var in checkboxes:
        var.set(0)  # Устанавливаем состояние чекбоксов из списка checkboxes в "не выбрано"


def threshold():
    global df_filtr
    # Запрашиваем номер стадии для изъятия полки
    print("Введите номер стадии для изъятия полки.")
    x1 = float(x1_entry.get())
    print("Введенный номер стадии:", x1)

     # Находим производные
    df_filtr['dy'] = np.gradient(df_filtr['Deviator Stress (kPa)'])
    # Фильтруем данные, оставляя только те строки, где z=1
    data_z1 = df_filtr[df_filtr['Stage Number'] == x1].copy() # Создаем копию

    # Выводим фильтрованные данные для отладки
    print("Фильтрованные данные:")
    print(data_z1)

   

    # Определяем порог для "полок"
    threshold = 2

    # Фильтруем данные, оставляя только те строки, где значение 'dy' больше или равно порогу
    data_cleaned = data_z1[np.abs(data_z1['dy']) >= threshold].copy()

    # Выводим очищенные данные для отладки
    print("Очищенные данные:")
    print(data_cleaned)

 
    # Создаем новый датафрейм, объединяя очищенные данные и данные без изменений
    df_filtr = pd.concat([data_cleaned, df_filtr[df_filtr['Stage Number'] != x1]])
    df_filtr = df_filtr.sort_index()
    print("Полка удалена.")
    plot_filter()

def treshold_ur():
    global df_filtr

    df_filtr.to_excel('2222222.xlsx', index=False)
    wb = load_workbook('2222222.xlsx')
    ws = wb.active

    chart = ScatterChart()
    chart.title = "График"
    chart.style = 2  # Используем стиль маркера

    chart.width = 21  # Ширина графика (в символах)
    chart.height = 15 # Высота графика (в символах)


    # Данные для первой серии (1-й столбец для X, 2-й столбец для Y)

    x_values = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
    y_values = Reference(ws, min_col=2, min_row=2, max_row=ws.max_row)

    series_1 = Series(y_values, xvalues=x_values, title_from_data=False)
 
    # Устанавливаем стиль маркера для точечной диаграммы
    series_1.marker.symbol = "circle"
    # не заливаем линию между маркерами (прозрачная)
    #series_1.graphicalProperties.line.noFill = True
    series_1.graphicalProperties.line.width = 10000  # Ширина линии (значение по умолчанию - 100000)



    # Добавляем серию данных в график
    chart.series.append(series_1)

    # Добавляем график в лист
    ws.add_chart(chart, "L2")  # Расположение графика


    # Сохраняем изменения в файле
    wb.save('2222222.xlsx')

    # Открываем файл Excel
    os.startfile('2222222.xlsx')

    # Ждем, пока файл закроется
    input("Press Enter after closing the Excel file...")
    df_updated = pd.read_excel('2222222.xlsx')
    df_filtr = df_updated
    print('Файл обновлен')
    # Удаляем файл
    os.remove('2222222.xlsx')
    print('Файл удален 222222222.xlsx')

def traxial():
    global df_filtr, df_triaxial, height
    df_triaxial = df_filtr

    if 'Back Volume (mm?)' in df_triaxial.columns:

        df_triaxial["epsilon_1 д.е"] = df_triaxial['Axial Displacement (mm)'] / height
        df_triaxial["epsilon_1"] = df_triaxial['Axial Displacement (mm)'] / height * 100     
        df_triaxial.loc["epsilon_1"] = df_triaxial['Axial Displacement (mm)'] / height * 100 #вычисляем столбец  
        first_row_volume = df_triaxial.iloc[0]['Back Volume (mm?)']    #фиксируем первую строку 
        print(first_row_volume)
        df_triaxial["epsilon_V"] = - (df_triaxial['Back Volume (mm?)'] - first_row_volume) / (1963.4953750 * height) #вычисляем объемную деформацию при диаметре образца в 50 мм
        Δhk = df_triaxial['Axial Displacement (mm)'].max()
        qmax = df_triaxial['Deviator Stress (kPa)'].max()
        Vi = (math.pi * 50 ** 2) / 4 * height 
        V = 196349.5408
        As = math.pi * (2.5 ** 2)
        h15_index = np.abs(df_triaxial['Axial Displacement (mm)'] - 15).idxmin() #находим индекс числа максимально приближенно к 15  
        Epsilon_H15 = df_triaxial.loc[h15_index, 'epsilon_V']# находим занчение эпсилон В в по индексу
        radialV = df_triaxial.iloc[0]['Radial Volume (mm?)']    #фиксируем первую строку 
        Ak = (V - (Epsilon_H15 * Vi)) / (height - Δhk) / 100
        Ac = (V - radialV) / height / 100
        b = (1 - (As / Ak) ) / (Δhk / height) 
        t = 0.36
        df_triaxial['Δσ1m, kPa'] = (41.65289256) * ((df_triaxial['epsilon_1'] * 0.01) + df_triaxial['epsilon_V'] / 3)
        df_triaxial['Δσ3m, kPa'] = (14.14736842 * df_triaxial['epsilon_V'])
        df_triaxial['Ai'] = Ac * (1 - df_triaxial['epsilon_V']) / (1 - b * (df_triaxial['epsilon_1'] * 0.01))
        df_triaxial['Correction Deviator Stress (kPa)'] = (df_triaxial['Load Cell (kN)'] / (df_triaxial['Ai'] * 0.0001)) - df_triaxial['Δσ1m, kPa'] - df_triaxial['Δσ3m, kPa']

        #print(df_triaxial['Correction Deviator Stress (kPa)'].head())
        #print(df_triaxial['Load Cell (kN)'].head())
        #print(df_triaxial['Ai'].head())
        #print(df_triaxial['Δσ1m, kPa'].head())
        #print(df_triaxial['Δσ3m, kPa'].head())
        #print(As, Ak, Ac, qmax, Vi, b, Δhk, Epsilon_H15)
    else:
        print("Столбец 'Back Volume (mm?)' отсутствует в DataFrame.")




def traxial_E():
    global df_triaxial, max_y, max_y_x, E50_natural, E, Ed
    # Предварительная обработка данных
    # Заменяем все значения NaN на 0 (или любое другое значение)
    

    # Ваши данные
    q11 = df_triaxial['Eff. Axial Stress (kPa)'].iloc[0]
    x11 = df_triaxial['epsilon_1'].iloc[0]
    q_1_6 = 1.6 * q11

    # Ваши данные
    x = df_triaxial['epsilon_1']
    q1 = df_triaxial['Eff. Axial Stress (kPa)']

   

  # Создаем объект функции интерполяции методом локальных полиномов
    interp_func = interp1d(q1, x, kind='next')

    # Интерполяция
    x_1_6 = interp_func(q_1_6)
    print(q11, x11)
    print(f"q_1_6: {q_1_6}, x_1_6: {x_1_6}")


    E = (q_1_6 - q11) / (x_1_6 - x11) * 0.1

   #создами маску для поиска и интерпаляции только в диапазоне, где load cell меняет свой знак.
     # Рассчитываем разницу между соседними значениями Load Cell
    d_load_cell = np.diff(df_triaxial['Load Cell (kN)'])

    # Ищем индекс, когда значение начинает уменьшаться
    end_index = None
    for i, diff in enumerate(d_load_cell):
        if diff < 0:
            end_index = i
            negative_count = 1  # Начинаем считать отрицательные значения с 1, так как мы уже нашли одно
            break

    # Если такой индекс не найден, то берем последний индекс
    if end_index is None:
        end_index = len(df_triaxial) - 1
    else:
        # Проверяем, идут ли отрицательные значения подряд (4 значения)
        for diff in d_load_cell[end_index + 1:]:
            if diff < 0:
                negative_count += 1
                if negative_count >= 7:
                    break
            else:
                # Если значение не отрицательное, обнуляем счетчик
                negative_count = 0
        else:
            # Если три или более отрицательных значений не найдены, то берем последний индекс
            end_index = len(df_triaxial) - 1


    # Создаем массивы значений только до этого индекса
    x_values = df_triaxial['epsilon_1'][:end_index]
    y_values = df_triaxial['Eff. Axial Stress (kPa)'][:end_index]
    print('индекс конечного элемента', end_index)
    print('значения массива для сортирорвки' , x_values, y_values)


       # Создаем DataFrame из массивов x и y
    df = pd.DataFrame({'x': x_values, 'y': y_values})

    # Удаляем дубликаты по значениям x и y
    df_unique = df.drop_duplicates()

        # Проверяем наличие пустых значений
    if df_unique.isnull().values.any():
        print("В датафрейме есть пустые значения.")
        # Удаляем строки с пустыми значениями
        df_unique = df_unique.dropna()

    # Получаем уникальные значения x и y
    unique_x_values = df_unique['x'].values
    unique_y_values = df_unique['y'].values

    print('значения уникального массива для сортирорвки' , unique_x_values, unique_y_values)

        # Удаляем дубликаты по значениям x
    unique_x_values, unique_indices_x = np.unique(unique_x_values, return_index=True)
    unique_y_values = unique_y_values[unique_indices_x]

    print('значения уникального массива для сортировки', unique_x_values, unique_y_values)


    try:
        # Интерполируем только эти значения методом локальных полиномов
        interp_func1 = interp1d(unique_y_values, unique_x_values, kind='slinear')

        # Производим интерполяцию для значения q_1_6
        x_1_6d = interp_func1(q_1_6)
    except ValueError as e:
        print("Ошибка:", e)
        print("Присваиваем x_1_6d значение x_1_6")
        x_1_6d = x_1_6
        
    print(f"q_1_6: {q_1_6}, x_1_6d: {x_1_6d}")

    # Вычисляем E
    Ed = (q_1_6 - q11) / (x_1_6d - x11) * 0.1
    print(x11, q11)

    E = (q_1_6 - q11) / (x_1_6 - x11) * 0.1  




    E_label.config(text=f"E: {E, Ed}")

 


    # Получение введенных значений
    #x1E = float(x1Е_entry.get())
    #x2E = float(x2Е_entry.get())
    #stageE_input = float(stageE_entry.get())


    x50 = df_triaxial['epsilon_1']
    y50 = df_triaxial['Correction Deviator Stress (kPa)']
    stage = df_triaxial['Stage Number']
    

      # Находим производные
    #dy = np.diff(y)
    # Находим индекс, где производная меняет знак
    #sign_change_index = np.where(np.diff(np.sign(dy)))[0][0]

    # Находим индекс первого значения y, которое больше значения в пункте 1
    #first_greater_index = np.argmax(y > y[sign_change_index])

    #if sign_change_index < len(y) and first_greater_index < len(y):
    #    mask1 = np.logical_and(sign_change_index <= np.arange(len(y)), np.arange(len(y)) <= first_greater_index)
    #    print(mask1)
    #else:
    #    print("Не удалось найти необходимые индексы.")
    # Определение Е50  
    # Найдем максимальное значение y
    max_y = np.max(y50)
    max_y_index = np.argmax(y50)  # Находим индекс максимального значения y
    max_y_x = x50.iloc[max_y_index]  # Получаем значение x, соответствующее максимальному значению y


    # Рассчитаем половину максимального значения y
    half_max_y = max_y / 2

    # Найдем индекс ближайшего значения x к половине максимального значения
    closest_index = np.argmin(np.abs(y50 - half_max_y))


    # Найдем значение y, которому соответствует найденный индекс
    x50_point = x50.iloc[closest_index]
    q50_point = y50.iloc[closest_index]
    print('qmax', max_y)
    print('q50',half_max_y)
    print('q50-natural', q50_point)
    print('epsilon50',x50_point)

    E50 = half_max_y / x50_point * 0.1
    E50_natural = q50_point / x50_point * 0.1

    E50_label.config(text=f"E50: {E50_natural}")

    print('x_y_max=', max_y_x)
    #df_triaxial.to_excel('2222222.xlsx', index=False)
     # Открываем файл Excel
    #os.startfile('2222222.xlsx')

    # Ждем, пока файл закроется
    #input("Press Enter after closing the Excel file...")
    #df_updated1 = pd.read_excel('2222222.xlsx')
    #df_triaxial = df_updated1
    #print('Файл обновлен')
    # Удаляем файл
    #os.remove('2222222.xlsx')
    #print('Файл удален 222222222.xlsx')
  
def traxial_Eur():
    global df_triaxial, x1ur, x2ur, y1ur, y2ur, Eur

    # Получить текст из виджетов
    x1E_text = x1Е_entry.get()
    x2E_text = x2Е_entry.get()

    # Разделить текст по пробелу и преобразовать в числа
    x1ur, y1ur = map(float, x1E_text.split())
    x2ur, y2ur = map(float, x2E_text.split())

    Eur = (y1ur - y2ur) / (x1ur - x2ur)
    Eur_label.config(text=f"Eur: {Eur}")
    print(x1E_text, x2E_text)
    print(Eur)

def plot_triaxial_q1_q3():
    global df_triaxial
    x = df_triaxial['epsilon_1'] #Присвоение x 
    y = df_triaxial['Correction Deviator Stress (kPa)'] # Присвоение у

    # Построение графика
   # Постройте график
    plt.figure()
    plt.plot(x, y, label='Данные', marker='o', color='b')

    plt.legend()

    # Добавьте заголовок и подписи к осям
    plt.title('график')
    plt.xlabel('Х')
    plt.ylabel('Y')
    # Добавление сетки
    plt.grid(True)
    # Отобразите график
    plt.show()

def plot_triaxial_V():
    global df_triaxial, anglesin, max_y_x, anglesin_degree
    # Создание данных для графика
    x = df_triaxial["epsilon_1 д.е"].dropna().values
    y = df_triaxial['epsilon_V'].dropna().values

    # Функция, обновляющая график при изменении положения точки
    def update_plot(event):
        global anglesin, anglesin_degree 
        if event.inaxes and event.xdata is not None:
            point = (event.xdata, np.interp(event.xdata, x, y))
            if min(x) <= point[0] <= max(x):
                plt.cla()
                plt.plot(x, y, label='График')
                plt.scatter(point[0], point[1], color='red', label='Выбранная точка')

                # Вычисляем угол дилатансии
                index = np.argmin(np.abs(x - point[0]))
                # Вычисляем разность между соседними значениями x и y
                delta_x = x[index + 10] - x[index - 10]
                delta_y = y[index + 10] - y[index - 10]
                deltae = delta_y / (delta_y - (2 * delta_x))
                angle = np.arctan2(delta_y, delta_x)  # Вычисляем угол касательной
                anglesin = np.arcsin(deltae)
                anglesin_degree = np.degrees(anglesin)
               # Вычисляем координаты второй точки на касательной для построения отрезка
                scale_factor = 0.03  # Коэффициент масштабирования
                tangent_x = [x[index] - scale_factor * np.cos(angle), x[index] + scale_factor * np.cos(angle)]
                tangent_y = [y[index] - scale_factor * np.sin(angle), y[index] + scale_factor * np.sin(angle)]

                #вычисляем координаты для линии с макс девиатором
                dev_x = [max_y_x * 0.01, max_y_x * 0.01]
                dev_y = [y.min(), y.max()]


                 
                # Построение касательной
                plt.plot(tangent_x, tangent_y, 'r--', label='Касательная')
                # добавляем гранинцу где наъодится qmax
                plt.plot(dev_x, dev_y, color='gray', alpha=0.5, linewidth=0.5, linestyle='--', label='макс девиатор')

                plt.title(f'График с касательной в точке, угол наклона: {anglesin_degree:.2f}°')
                plt.xlabel('X')
                plt.ylabel('Y')
                plt.legend()
                plt.grid(True)
                plt.draw()
                plt.gca().invert_yaxis()  # Переворачиваем ось y

                print("delta_x = ", delta_x, "delta_y=", delta_y)
                print("angle в радианах:", angle, "angle в градусах:", np.degrees(angle))
                print("angle в радианах sin:", anglesin, "angle в градусах sin:", np.degrees(anglesin))
                Dila_label.config(text=f"Угло дилатансии и q max: {np.degrees(anglesin)} {max_y}")

            else:
                print("Точка находится за пределами допустимого диапазона x")
        else:
            print("Событие не в области осей или xdata не определен")

    # Создание начального графика
    plt.figure(figsize=(8, 6))
    plt.plot(x, y, label='График')
    plt.title('График с касательной в точке')
    plt.xlabel('X')
    plt.ylabel('Y')
    plt.legend()
    plt.grid(True)
    plt.gca().invert_yaxis()  # Переворачиваем ось y

    # Привязка обработчика событий к графику
    plt.gcf().canvas.mpl_connect('button_press_event', update_plot)

    # Отображение графика
    plt.show()
    return(anglesin, anglesin_degree)

def import_file_gds():
    global df
    file_path = filedialog.askopenfilename()
    # Загрузить данные (возможно используется кодтровака, и нужно указать другой разделитель между данными df = pd.read_csv('sinusoida.csv', encoding='utf-8', quotechar='"')
    print("Загрузка данных...")
    source_file = file_path
    # Желаемое новое расширение (например, .csv)
    new_extension = '.csv'

    # Изменяем расширение файла
    base_name, _ = os.path.splitext(source_file)
    new_file = base_name + new_extension

    # Копируем исходный файл с новым именем и расширением
    shutil.copy(source_file, new_file)
    # Теперь у вас есть файл с новым расширением (new_file)

    try:
        with open(new_file, 'r') as file:
            lines = file.readlines()
    except FileNotFoundError:
        print(f"Файл {new_file} не найден.")
        exit(1)

    # Удаление последней строки (если она есть)
    if len(lines) > 1:
        lines = lines[:-1]

    # Запись обновленных строк обратно в файл CSV
    with open(new_file, 'w') as file:
        file.writelines(lines)


    # Например, для чтения нового файла:
    with open(new_file, 'r') as file:
        data = pd.read_csv(new_file, skiprows=57, skipfooter=3)
        data = data[['Axial Displacement (mm)', 'Deviator Stress (kPa)', 'Back Volume (mm?)','Eff. Axial Stress (kPa)', 'Load Cell (kN)', 'Radial Volume (mm?)', 'Time since start of test (s)', 'Time since start of stage (s)', 'Stage Number' ]] #загрузил толь эти столбцы   
    df = data

    print("Данные загружены.")
    print(df)


def import_file_geo():
    global df
    file_path = filedialog.askopenfilename()
    # Загрузить данные (возможно используется кодтровака, и нужно указать другой разделитель между данными df = pd.read_csv('sinusoida.csv', encoding='utf-8', quotechar='"')
    print("Загрузка данных...")
    source_file = file_path

    data = pd.read_csv(source_file, sep='\t', decimal=',')
    data = data[['ChVerticalDeformation_mm', 'ChDeviator_kPa', 'ChPoreVolumeBottom_cm3', 'ChVerticalEffectivePress_kPa', 'ChVerticalLoadInner_N', 'ChRadialDeformation_mm', 'Time', 'StageTime', 'StageId']]
    data = data.rename(columns={'ChVerticalDeformation_mm': 'Axial Displacement (mm)', 'ChDeviator_kPa': 'Deviator Stress (kPa)', 'ChPoreVolumeBottom_cm3': 'Back Volume (mm?)', 'ChVerticalEffectivePress_kPa': 'Eff. Axial Stress (kPa)', 'ChVerticalLoadInner_N': 'Load Cell (kN)', 'ChRadialDeformation_mm': 'Radial Volume (mm?)', 'Time': 'Time since start of test (s)', 'StageTime': 'Time since start of stage (s)', 'StageId': 'Stage Number' })  #переименовал столбцы
    df = data

    

    print("Данные загружены.")

def save1():
    global df # Объявление, что вы используете глобальную переменную df
    if df is not None:
        save_file_path = filedialog.asksaveasfilename(defaultextension=".csv", filetypes=[("csv_files", "*.csv"), ("All files", "*.*")])
        if save_file_path:
            df.to_csv(save_file_path, index=False)
            print("Данные сохранены в файл:", save_file_path)
            update_file_list()
        else:
            print("Сохранение отменено")
    else:
        print("Нет данных для сохранения")

def save2():
    global df_triaxial, height, max_y,  E50_natural, E, Ed, x1ur, x2ur, y1ur, y2ur, Eur, anglesin, anglesin_degree
    
    # Предложить пользователю сохранить результаты в файл Excel
    save_file_path = filedialog.asksaveasfilename(defaultextension=".xlsx",
                                                  filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")])
    if save_file_path:
    # Сохранение данных в Excel
        df_triaxial.to_excel(save_file_path, index=False)
 

    # Загружаем существующий Excel-файл
        wb = load_workbook(save_file_path)
        ws = wb.active

              # Удаление столбцов, содержащих ".1" в имени
        columns_to_drop = [column for column in ws.iter_cols(1, ws.max_column) if ".1" in column[0].value]
        for column in columns_to_drop:
            ws.delete_cols(column[0].column)


         # Добавление заголовков
        ws.append(['Global Variables'])
        ws.append(['Variable', 'Value'])
        
        # Добавление глобальных переменных
        variables = {
            'Высота образцы': height,
            'максимальный девиатор': max_y,
            'E50': E50_natural,
            'E': E,
            'Ed': Ed,
            'x1ur': x1ur,
            'x2ur': x2ur,
            'y1ur': y1ur,
            'y2ur': y2ur,
            'Eur': Eur,
            'угол дилатансии': anglesin_degree
        }
        
             # Получение индекса столбца
        row_number = 3  # Начальный индекс строки
        column_index = 18  # Индекс столбца R

        for variable, value in variables.items():
            ws.cell(row=row_number, column=column_index, value=variable)
            ws.cell(row=row_number, column=column_index + 1, value=value)
            row_number += 1  # Переходим к следующей строке


        chart = ScatterChart()
        chart.title = "График зависимости девиатора напряжений от относительной деформации"
        chart.x_axis.title = "Относительная деформация, \u03B5 1 %."  # Название оси X
        chart.y_axis.title = "Девиатор напряжений, КПа."  # Название оси Y
        chart.style = 2  # Используем стиль маркера

        chart.width = 21  # Ширина графика (в символах)
        chart.height = 15 # Высота графика (в символах)


        # Данные для первой серии (1-й столбец для X, 2-й столбец для Y)

        x_values = Reference(ws, min_col=11, min_row=2, max_row=ws.max_row)
        y_values = Reference(ws, min_col=16, min_row=2, max_row=ws.max_row)

        series_1 = Series(y_values, xvalues=x_values, title_from_data=False)
     
        # Устанавливаем стиль маркера для точечной диаграммы
        series_1.marker.symbol = "circle"
        # не заливаем линию между маркерами (прозрачная)
        #series_1.graphicalProperties.line.noFill = True
        series_1.graphicalProperties.line.width = 10000  # Ширина линии (значение по умолчанию - 100000)



        # Добавляем серию данных в график
        chart.series.append(series_1)

        # Добавляем график в лист
        ws.add_chart(chart, "V2")  # Расположение графика


 
        chart2 = ScatterChart()
        chart2.title = "График зависимости объемной деформации от относительной"
        chart2.x_axis.title = "Относительная деформация, \u03B5 1 доли ед."  # Название оси X
        chart2.y_axis.title = "Объемная деформация, \u03B5 v доли ед."  # Название оси Y
        chart2.style = 2  # Используем стиль маркера

        chart2.width = 21  # Ширина графика (в символах)
        chart2.height = 15 # Высота графика (в символах)


        # Данные для первой серии (1-й столбец для X, 2-й столбец для Y)

        x_values2 = Reference(ws, min_col=10, min_row=2, max_row=ws.max_row)
        y_values2 = Reference(ws, min_col=12, min_row=2, max_row=ws.max_row)

        series_2 = Series(y_values2, xvalues=x_values2, title_from_data=False)
     
        # Устанавливаем стиль маркера для точечной диаграммы
        series_2.marker.symbol = "circle"
        # Настраиваем параметры оси y
        chart2.y_axis.scaling.orientation = "maxMin"  # Переворачиваем ось y
        # не заливаем линию между маркерами (прозрачная)
        #series_1.graphicalProperties.line.noFill = True
        series_2.graphicalProperties.line.width = 10000  # Ширина линии (значение по умолчанию - 100000)



        # Добавляем серию данных в график
        chart2.series.append(series_2)

        # Добавляем график в лист
        ws.add_chart(chart2, "V32")  # Расположение графика


 
        # Сохраняем изменения в файле
        wb.save(save_file_path)
        print("Данные сохранены в файл:", save_file_path)
    else:
        print("Сохранение отменено")

def import_and_save_gds():
    import_file_gds()
    save1()

def import_and_save_geo():
    import_file_geo()
    save1()





# Создание главного окна
root = tk.Tk()
root.title("Trixial v.0.3")

# Создание холста
canvas = tk.Canvas(root)
canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

# Создание фрейма для размещения виджетов
frame = tk.Frame(canvas)
frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

# Привязка виджетов к холсту
canvas.create_window((0, 0), window=frame, anchor='nw')

# Создание вертикальной прокрутки для холста
scrollbar = ttk.Scrollbar(root, orient="vertical", command=canvas.yview)
scrollbar.pack(side=tk.RIGHT, fill='y')

# Привязка прокрутки к холсту
canvas.config(yscrollcommand=scrollbar.set)
canvas.bind('<Configure>', lambda e: canvas.configure(scrollregion=canvas.bbox("all")))




# Рамка для списка файлов
frame_files = ttk.Frame(root)
frame_files.pack(side=tk.LEFT, fill=tk.Y, expand=False)

# Список файлов
file_listbox = tk.Listbox(frame_files)
file_listbox.pack(side="left", fill="both", expand=True)

# Добавление прокрутки для списка файлов
scrollbar = ttk.Scrollbar(frame_files, orient="vertical", command=file_listbox.yview)
scrollbar.pack(side="right", fill="y")
file_listbox.config(yscrollcommand=scrollbar.set)

directory_path = "C:/Users/Егор/Desktop/python/file1"

update_thread = threading.Thread(target=update_file_list)
update_thread.daemon = True  # Поток завершится при завершении основной программы
update_thread.start()

# Привязка функции к событию выбора файла
file_listbox.bind('<<ListboxSelect>>', on_file_select)
# 

# Создание метки для вывода названия файла
Name_file = tk.Label(root, text="")
Name_file.pack()

# Добавление кнопки для сохранения выбранных данных
save_button = tk.Button(root, text="Показать график согласно выбранным стадиям", command=save_filtered_data)
save_button.pack()

# Создание чекбокса "Выделить выбранные"
select_from_list_checkbox = tk.Checkbutton(root, text="Выделить выбранные", command=select_from_list)
select_from_list_checkbox.pack()

# Создание чекбокса "Снять выделение с выбранных"
deselect_from_list_checkbox = tk.Checkbutton(root, text="Снять выделение с выбранных", command=deselect_from_list)
deselect_from_list_checkbox.pack()

#Создание кнопки обнулить перемещение 
reset_button = tk.Button(root, text="Обнулить перемещение", command=reset_displacement)
reset_button.pack()

#Создание кнопки обнулить девиатор
reset_button = tk.Button(root, text="Обнулить девиатор", command=reset_deviator)
reset_button.pack()

# Виджеты для ввода стадии для убирания полки
x1_label = tk.Label(root, text="Введите номер стадии для изъятия полки:")
x1_label.pack()
x1_entry = tk.Entry(root)
x1_entry.pack()


#Создание кнопки удалить редакиторавть график
reset_button = tk.Button(root, text="Редактировать график", command=treshold_ur)
reset_button.pack()

#Создание кнопки удалить полку
reset_button = tk.Button(root, text="Удалить полку", command=threshold)
reset_button.pack()

#Создание кнопки вычислить данные для основынх расчетов
reset_button = tk.Button(root, text="Вычислить основные дарнные", command=traxial)
reset_button.pack()

#Создание кнопки потсроить график по девиатору
reset_button = tk.Button(root, text="Построить основной график", command=plot_triaxial_q1_q3)
reset_button.pack()

#Создание кнопки потсроить график по деформациям
reset_button = tk.Button(root, text="Построить график деформаций", command=plot_triaxial_V)
reset_button.pack()

#Виджеты для определенеия модуля деформации

# Виджеты для ввода ху для В и А
x1Е_label = tk.Label(root, text="Введите xy точки В (ху разделены пробелом):")
x1Е_label.pack()
x1Е_entry = tk.Entry(root)
x1Е_entry.pack()

x2Е_label = tk.Label(root, text="Введите xy точки A:")
x2Е_label.pack()
x2Е_entry = tk.Entry(root)
x2Е_entry.pack()

#Создание кнопки расчитать Еur
reset_button = tk.Button(root, text="Расчитать Еur", command=traxial_Eur)
reset_button.pack()
# Создание метки для вывода Еur 
Eur_label = tk.Label(root, text="")
Eur_label.pack()


# Создание метки для вывода Е 
E_label = tk.Label(root, text="")
E_label.pack()

#Создание кнопки расчитать модуль деформации Е
reset_button = tk.Button(root, text="Расчитать Е", command=traxial_E)
reset_button.pack()

# Создание метки для вывода E50
E50_label = tk.Label(root, text="")
E50_label.pack()

# Создание метки для вывода угла дилатансии и max q
Dila_label = tk.Label(root, text="")
Dila_label.pack()

# Кнопка для сохранения файла Эксель
save_to_exel_button = tk.Button(root, text="Сохранить в формате Excel", command=save2)
save_to_exel_button.pack()

# Создание меню
menu_bar = tk.Menu(root)
root.config(menu=menu_bar)

file_menu = tk.Menu(menu_bar, tearoff=0)
menu_bar.add_cascade(label="Файл", menu=file_menu)

file_menu.add_command(label="Импортировать GDS", command=import_and_save_gds)
file_menu.add_command(label="Импортировать Геотек", command=import_and_save_geo)
file_menu.add_separator()
file_menu.add_command(label="Выход", command=root.quit)






# Запуск главного цикла
root.mainloop()
